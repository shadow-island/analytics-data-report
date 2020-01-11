# -*- coding: utf-8 -*-
# python 3X
'''
본 App 장점:
    Youtube보다 offline이 좋은점: internet안될때도되고, 일단 곡별로 가중치를 줄수있음
Next:
    git연동완료,email이 중요
    scroll이 불편해: print time 보강작업:프로그램 재실행시도 보기 가능하니, 일단 메모리후 file화로
        1 print time을 강화?=>
        2 또는 체크타임 결과를 어디에 저장?
    0. 궁극적으로는 c# window style
    실행횟수? force new?    
Release note
#기능
    media파일만 하기
#UI
    입력 숫자 error handling, 메뉴표시하자,마지막 status표시
    version_info

call func 즉 4줄이상이 2번나올때(총8줄) refactoring하면 옳다
call func
def func
    sentence 1
    sentence 2
    sentence 3
    return sentence 4
'''

import os
import sys
import datetime

def read_file(file_name):
    return_flag = True
    return_data = None
            
    if os.path.exists(file_name):
        #euk
        my_file = open(file_name,'r')
        return_data  = my_file.readlines()
        my_file.close()
    else:
        return_flag = False
                            
    return (return_flag, return_data)

def total_sec_2_readable(total_secs):
    text = ''
    remained_sec = total_secs
    h               = int(remained_sec / 3600)    
    if h >= 24:
        d = (int)(h / 24)
        text = str(d) + '일'
        h = h % 24 
    
    remained_sec    = total_secs % 3600         
    print_mi        = int(remained_sec / 60)
    remained_sec    = remained_sec % 60

    if h > 0:
        text += str(h)+'시간 '
    
    if print_mi > 0:
        text += str(print_mi) + '분 '
         
    text += str(remained_sec) + '초'
    
    return text

def get_last(local_file_name, bar):
    # File이 없는경우 대비
    #datetime_old_time   = datetime.datetime(2014, 8, 3, 0,0,0)
    datetime_old_time   = datetime.datetime.now()
    int_old_sec_gap     = 0

    (return_flag, data_list)   = read_file(local_file_name)

    if return_flag == True:
        for line in data_list:
            #item_list값을 밖에서 쓸수없음 bc)file이 없을경우가 잇으므로
            item_list = line.split(bar)
            s_old   = item_list[0]

            if s_old[:4].isdigit():
                print('Proper File Format')
                #code 정리
                y       = int(s_old[:4])
                mo      = int(s_old[5:7])
                d       = int(s_old[8:10])
                h       = int(s_old[11:13])
                minutes = int(s_old[14:16])
                s       = int(s_old[17:19])
                datetime_old_time = datetime.datetime(y,mo,d,h,minutes,s)
                int_old_sec_gap   = int(item_list[1])
    else:
        print ('file error')
    #file reading end
    return (datetime_old_time, int_old_sec_gap)

def common_print_next(int_new_gap, datetime_old_time, int_old_sec_gap):
    print ('이번 듣기 주기: ' + total_sec_2_readable(int_new_gap))
    
    print ('\n다음 평균 주기: ' + total_sec_2_readable(int_old_sec_gap))
    next_gap_timedelta = datetime.timedelta(0,int_old_sec_gap)
    print ('다음 목표 시간: ' + str(datetime_old_time + next_gap_timedelta) + '\n')

def check_update():
    print('menu_song_time_check')
    import datetime
    
    # for both read and save
    local_file_name = 'eukM2.log'
    bar             = '~'
    #~
    
    (datetime_old_time, int_old_sec_gap) =  get_last(local_file_name, bar)
    
    print ("이전 체크 시간: " + str(datetime_old_time))
    print ('이전 평균 주기:', total_sec_2_readable(int_old_sec_gap))
    
    datetime_new        =   datetime.datetime.now()  
    str_date            =   str(datetime_new)[:4]       + "-" +  \
                            str(datetime_new)[5:7]      + "-" +  \
                            str(datetime_new)[8:10]     + " " +  \
                            str(datetime_new)[11:13]    + ":" +  \
                            str(datetime_new)[14:16]    + ":" +  \
                            str(datetime_new)[17:19]

    datetime_gap_new    =   datetime_new - datetime_old_time
    
    print ("\n현재 체크 시간: " + str_date)
    
    int_new_gap = int(datetime_gap_new.days * 60*60*24 + (datetime_new - datetime_old_time).seconds)
    int_old_sec_gap = int( (int_new_gap + int_old_sec_gap)/2 )
        
    common_print_next(int_new_gap, datetime_old_time, int_old_sec_gap)
    
    if (int_old_sec_gap <= int_new_gap):
        b_text = 'increased'
        print ("천천히 들으셨네요. you can still enjoy them, change priority => go 4")
        '''
        command = '그래도추가!'
        start_n = 1
        end_n   = 6
        ran_n = random_range_until(start_n, end_n)    
        if (ran_n == 1):
            print('%s %d~%d => %d' %(command, start_n, end_n, ran_n))
            input(command)
            return
        else:
        '''
        print ("종료!")
    else:
        b_text = 'reduced'
        print ("이번 듣기 주기가 짧았네요.자꾸 들으니 지겨우시죠. 바야흐로 new music 추가 타이밍!")        
    
    print('ok?')
        
    my_file = open(local_file_name,"w")    
    my_file.write(
        str_date + bar +
        str(int_old_sec_gap) + bar +
        str(int_old_sec_gap) + bar +
        b_text )          
    my_file.close()    
    
    if b_text == 'reduced':
        pass #input('ready?')
    return int_new_gap


def create_db():
    #여기 두는것은 전체 플로그램을 그냥 시간체크용으로만 쓰는경우도 있어서
    import openpyxl
    from openpyxl import Workbook
        
    old_dict = {}
    if (os.path.exists(db_xls)):
        wb = openpyxl.load_workbook(db_xls)
        sheet_list = wb.get_sheet_names()
        #print (sheet_list)
        sheet = wb.get_sheet_by_name(sheet_list[0])        
        for i in range(1,sheet.max_row + 1):
            key = sheet.cell(row = i, column = 1).value     
            b_debug = True
            b_debug = False        
            if (b_debug == True):
                print (key + ':' + str(sheet.cell(row = i, column = 2).value))
            old_dict[key] = sheet.cell(row = i, column = 2).value
    #endif    

    print ('[creating DB]')
    #
    if os.path.exists(db_xls):
        if os.path.exists('backup_db.xlsx'):
            os.remove('backup_db.xlsx')
        try:    
            os.rename(db_xls,'backup_db.xlsx')
        except :
            print ('e')
            return False
                
    
    file_path_name_list = []
    
    for root, dirs, files in os.walk(mp3_folder_no_endslash):
        for file in files:
            text = os.path.join(root, file)    
            def print_and_append(text):
                try:                
                    print(text)
                except:
                    pass                
                file_path_name_list.append(text)
                      
            if text.find('mp4') != -1 or text.find('mp3') != -1 or text.find('avi') != -1 or text.find('webm') != -1  or text.find('mkv') != -1 or text.find('wav') != -1:
                print_and_append(text)
                
    i_len = len(file_path_name_list)
    print  ('file_path_name_list len:' + str(i_len))
    #for linux
    file_path_name_list.sort();     

    wb = Workbook()
    sheet = wb.active

    row_num = 0;
    for file_path_name in file_path_name_list:
        key = file_path_name	
        row_num += 1
        sheet.cell(row=row_num,column=1).value=key
        
	#for k in old_dict.keys():
	#print 'keys~'
	#print old_dict.keys()
        if key in old_dict.keys():
            sheet.cell(row = row_num,column = 2).value = old_dict[file_path_name]
        else:
            sheet.cell(row = row_num,column = 2).value = '1'
        #print file_path_name+ ':' + str(sheet.cell(row = row_num,column = 2).value)
                
    print ('exist:row_num:' + str(row_num))                
    wb.save(filename = db_xls)   
    return True
    
def create_db_N_create_selected_playlist():
    # create_db 자체를 실패하면 작업취소
    if create_db() == False:        
        return
            
    create_selected_playlist()
    return 

def create_selected_playlist():
    import openpyxl
    from openpyxl import Workbook
    #쓸때 필요
    #pre location for error
    import random
    import shutil
    
    wb = openpyxl.load_workbook(db_xls)    

    #txt_file = open("eukM2.m3u","w")
    txt_file = open("eukM2.m3u","w",encoding="utf-8")
    
    count = 0
    
    sheet_list = wb.get_sheet_names()
    #print sheet_list
    sheet = wb.get_sheet_by_name(sheet_list[0])
        
    for i in range(1,sheet.max_row + 1):
        i_value = int(sheet.cell(row = i, column = 2).value)                

        print_text = sheet.cell(row = i, column = 1).value + ':' + str(sheet.cell(row = i, column = 2).value)
        random_num = random.randrange(1,i_value + 1)
        if (random_num == 1):                         
            try:
                print (str(i_value) + '-' + str(random_num) + ':' + print_text + ' Ok')                 
            except:
                pass
            
            out_txt = sheet.cell(row = i, column = 1).value + '\n'
            if b_windows_or_linux == False:                
                out_txt = out_txt.replace(mp3_folder_no_endslash + '\\','/my/ongaku/')
                out_txt = out_txt.replace('\\','/')
            txt_file.write(out_txt)
            count = count + 1
        else:
            try :
                print (print_text + ' X')                
            except:
                pass
        
    print (str(count) + '/' + str(sheet.max_row))  
        
    #복사하기전에 close
    txt_file.close()    
    if b_windows_or_linux == False:
        shutil.copy('eukM2.m3u', '\\\\VBOXSVR\\my\\')
        print('복사')

def print_time():
    # for both read and save    
    local_file_name = 'eukM2.log'
    bar             = '~'
    (datetime_old_time, int_old_sec_gap) =  get_last(local_file_name, bar)

    common_print_next(int_new_gap, datetime_old_time, int_old_sec_gap)
    #~
    return

##main    
'''
main
'''
if __name__ == '__main__':
    ##
    import codecs
    
    work = 563
    work = round(work/60 * 1.1,1)
    
    #source file 이름찾기?
    source_file = codecs.open('yutjin_music.py', 'r','utf-8')
    line_list   = source_file.readlines()    
    source_file.close()
    
    count = 0
    for _ in line_list:
        count += 1
    print('Yutjin Music Version: ' + str(work) + 'H ' + str(count) + '\n')
    
    # Python Version
    def find_num(text,finding_text):
        index       = text.find(finding_text)
        size        = len(finding_text)
        location    = index + size
        return text[location:location + 1]    

    major = find_num(str(sys.version_info),'major=')
    minor = find_num(str(sys.version_info),'minor=')
    micro = find_num(str(sys.version_info),'micro=')
    print ('My Python Version: ' +  major + '.' + minor  + '.' + micro)
    #~

    ## global
    b_windows_or_linux = True
    db_xls = 'db.xlsx'    

    # 무슨 모드인지 따라 folder 정함
    if os.path.exists('euk_music_2p3.linux'):
        b_windows_or_linux = False
        print ('Office Mode')
    else:
        b_windows_or_linux = True
        print ('Home windows Mode')
    if b_windows_or_linux:
        mp3_folder_no_endslash = 'D:\\my\\public_data\\music\\0.rank'
    else:    
        mp3_folder_no_endslash = '\\\\VBOXSVR\\my\\ongaku'
    #~
    ##~ end of global

    #init
    int_new_gap = 0
    while True:            
        function_map = {            
            1:check_update,
            2:create_db,
            3:create_db_N_create_selected_playlist,
            4:create_selected_playlist,
            5:print_time
        }

        print ('[Menu]')
        print ('''
            1:check time
            2:After new songs create, create new DB 
            3:After new songs,and create_selected_playlist
            4:Create new playlist
            5:print_time
            ''')
        
        #숫자를 받아들일때까지 물음
        while True:
            input_v = input('?')            
            if input_v.isdigit() == False:
                continue
            else:
                input_int = int(input_v)
                if input_int < 0 or 5 < input_int:
                    continue
                break
                
        if input_int == 1:
            int_new_gap = check_update()            
            print('(int_new_gap)',int_new_gap)
        elif input_int == 0:
            quit()
        else:
            function_map.get(input_int)()
    
    print ('End')
