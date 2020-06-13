
# -*- coding: utf-8 -*-
# python 3X
'''

'''

def runSubs():
    #왼쪽모니터사용(세로수많음_)
    is_right = True
    #is_right = False
    #is_sub_sort = True
    is_sub_sort = False #조회수
    
    (key, data) = load(is_right)
    
    is_debug_mode = True
    is_debug_mode = False

    #글꼴 38로 했을때 7개 까지 나옴 ,너비61
    total = len(data)
    out_table = []
    print('Getting real time data')
    for i in range(total -1 ,-1, -1):
        value = data[i][1]
        le = len(value)
        #print(le)
        if le <= 17:
            channel_data = urllib.request.urlopen("https://www.googleapis.com/youtube/v3/channels?part=statistics&forUsername="+value+"&key="+key).read()
            snippet = urllib.request.urlopen("https://www.googleapis.com/youtube/v3/channels?part=snippet&forUsername="+value+"&key="+key).read()
            
        else:
            channel_data = urllib.request.urlopen("https://www.googleapis.com/youtube/v3/channels?part=statistics&id=" + value + "&key="+key).read()             
            snippet = urllib.request.urlopen("https://www.googleapis.com/youtube/v3/channels?part=snippet&id="+value+"&key="+key).read()
            
        if is_debug_mode == True:
            print(json.loads(channel_data)["items"]) #for debug
            
        #print(json.loads(channel_data))
        #print(json.loads(snippet))
        
        subs = int(json.loads(channel_data)["items"][0]["statistics"]["subscriberCount"])        
        view = json.loads(channel_data)["items"][0]["statistics"]["viewCount"]
        date = json.loads(snippet)["items"][0]["snippet"]["publishedAt"]
        def get_date(datetime_value):
            txt = str(datetime_value)
            index = txt.find('T')
            txt = txt[:index]    
            return txt
                    
        one_record = []
        one_record.append(data[i][0])
        one_record.append(subs)
        one_record.append(int(view))
        one_record.append(get_date(date))        
        
        out_table.append(one_record)
        #break
    
    print()
    print()
    print()
    print()
    print()
    print()
    print()
    print()
    print()
    print()
    print()
    print()    
    print()    
    print()    
    print()    
    print()    
    print()    
    print()    
    print()    
    print()    
    print()    
    print()
    print()
    print()
    print()
    print()
    print()
    print()
    print()
    print()
    print()
    print()
    print()
    
    print(' Real Time Data!')
    
    import datetime
    def time2text(datetime_value):
            txt = str(datetime_value)
            index = txt.find('.')
            return txt[:index]
    #print("실시간:" + time2text(datetime.datetime.now()))
    print()
    #out_table = sorted(out_table, key=lambda item: item[1], reverse=True) #online?
    
    is_reverse = False
    if is_debug_mode == True:
        is_reverse = True 
    # reverse는 감소방향 true는 감소방향(print용) false는 증가방향(real time용)
    
    
    
    if is_sub_sort == True:
        #out_table = sorted(out_table, key=lambda item: item[1], reverse = False)       
        out_table = sorted(out_table, key=lambda item: item[1], reverse=is_reverse)
    else:
        out_table = sorted(out_table, key=lambda item: item[2], reverse=is_reverse)        
    
    total = len(out_table)
    import time
    for i in range(0,len(out_table)):        
        if is_debug_mode != True:            
            time.sleep(2)    
        
        name = str(out_table[i][0])
        size = len(name)
        if size <= 3:
            column = '\t\t\t\t'
        else:
            column = '\t'
            
        if is_debug_mode == True:
            num = i + 1
        else:
            num = total - i
            
        print(' ' + str(num) + '위:\t' + name)
        
        if is_sub_sort == True:
            print(' \t구독자:' + str(format(out_table[i][1], ',')) +' 총조회수:' +  str(format(out_table[i][2], ','))  +' 개설:' + out_table[i][3])            
        else:            
            print(' \t' + '총조회수:' +  str(format(out_table[i][2], ',')) + ' 구독자:' + str(format(out_table[i][1], ',')) +' 개설:' + out_table[i][3])
        print()
    print()    
    time.sleep(5)
    return

def runList(list_id):
    
    (key,data) = load(None)
    while(True):    
        base_url = 'https://www.googleapis.com/youtube/v3/playlistItems?part=snippet&maxResults=50&playlistId=' + list_id + '&key=' + key
        nextPageToken = ""

        total_n = 1
        import datetime
        datetime_total = datetime.datetime(1,1,1,0,0,0)
        part_mode = False
        for n in range(3):
            is_breaker = False
            this_url = base_url + "&pageToken=" + nextPageToken
            json_data = urllib.request.urlopen(this_url).read()     

            item_data = json.loads(json_data)["items"]
            print(len(item_data))
            for i in range(len(item_data)):
                item_data = json.loads(json_data)["items"][i]['snippet']['title']
                print(str(total_n) + ":"+ item_data)        
                videoId = json.loads(json_data)["items"][i]['snippet']['resourceId']['videoId']
                print(videoId)
                         
                v_json_data = urllib.request.urlopen("https://www.googleapis.com/youtube/v3/videos?id=" + str(videoId) + "&part=contentDetails&key=" + key).read()
                v_item_data = json.loads(v_json_data)["items"][0]["contentDetails"]['duration']
                
                v_item_data = v_item_data.replace('PT','')        
                #debug 
                print(v_item_data)
                h = m = s = 0
                i = v_item_data.find('H')
                if i != -1:
                    h = v_item_data[:i]
                print(h)
                v_item_data = v_item_data[i+1:]
                
                i = v_item_data.find('M')
                if i != -1:
                    m = v_item_data[:i]
                print(m)
                v_item_data = v_item_data[i+1:]
                
                i = v_item_data.find('S')
                if i != -1:
                    s = v_item_data[:i]
                print(s)
                
                date_time_obj = datetime.datetime(1,1,1,int(h),int(m),int(s))
            
                datetime_now = datetime.timedelta(hours = date_time_obj.hour, minutes = date_time_obj.minute, seconds = date_time_obj.second)
                
                datetime_total += datetime_now
                print(datetime_now)           
                print(str(datetime_total.hour) + ":" + str(datetime_total.minute))
                print()
                
                if part_mode == True and total_n == 17:
                    print(datetime_total)
                    is_breaker = True
                    break                
                total_n += 1
                
            if is_breaker == True:
                break
            
            if "nextPageToken" in json.loads(json_data):
                nextPageToken = json.loads(json_data)["nextPageToken"]
                print(nextPageToken)
                n += 1
            else:
                break
        print(datetime_total)
        input("do you want to check this list again?")
    return

def load(is_right):
    db_xls = "db.xlsx"

    import openpyxl
    from openpyxl import Workbook
    import os

    data = []

    if os.path.isfile(db_xls) == True:                
        wb = openpyxl.load_workbook(db_xls)
        sheet_list = wb.get_sheet_names()
        #print sheet_list
        sheet = wb.get_sheet_by_name('k')
        s_value = sheet.cell(row = 1, column = 1).value
        print(s_value)
        key = s_value
        
        if is_right == True:
            sheet = wb.get_sheet_by_name('c')
        else:
            sheet = wb.get_sheet_by_name('c_l')
        
        for i in range(1,sheet.max_row + 1):
            record = []
            record.append(sheet.cell(row = i, column = 1).value)
            record.append(sheet.cell(row = i, column = 2).value)
            data.append(record)
    else:
        print(db_xls + " 없음")    
    
    return (key, data)
    
#main-----------------------------------------
#최소한 2020-02-14 이전에 개발 
work = 127
work = round(work/60 * 1.1,1)
#source file 이름찾기?
import codecs
source_file = codecs.open('yutjin_yutu.py', 'r','utf-8')
line_list   = source_file.readlines()    
source_file.close()

count = 0
for _ in line_list:
    count += 1
print('Yutjin Youtube Version 2: ' + str(work) + 'H ' + str(count) + '\n')

# Python Version
def find_num(text,finding_text):
    index       = text.find(finding_text)
    size        = len(finding_text)
    location    = index + size
    return text[location:location + 1]    
import sys
major = find_num(str(sys.version_info),'major=')
minor = find_num(str(sys.version_info),'minor=')
micro = find_num(str(sys.version_info),'micro=')
print ('My Python Version: ' +  major + '.' + minor  + '.' + micro)
#~


import urllib.request
import json

is_subs_mode = False
import sys
if len(sys.argv) == 1:        
    #print('normmal mode:no argv')
    print('1)list\n')
    print('2)channel sub\n')
    input_menu = input('?')
    
    if input_menu == '1' or len(input_menu) > 1:
        
        list_id = 'UUClVppyt5FlY8rCTLGDgOIA'
        list_id = 'UC0Fq24M32ruKPcMH2xxxxxx' # UC로 시작하면 채널명임
        #list_id = 'PL2efNl7MkFICURkXwsmymaFGg-NIwzj49'
        
        if len(input_menu) == 1:
            list_id = input('list?')
        else:
            list_id = input_menu
        runList(list_id)
    else: #enter도 가능
       runSubs()
else:
    is_subs_mode = True
    #sys.argv
    print(sys.argv[1])    
    #runSubs()

    
exit(0)
