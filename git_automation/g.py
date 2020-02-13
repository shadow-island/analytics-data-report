# -*- coding: utf-8 -*-
# python 3X
'''
본 App 장점:
Next:
Release note
#기능
#UI
'''

import os
import sys
import datetime

def read_file(file_name):
    return_flag = True
    return_data = None
            
    if os.path.exists(file_name):
        #euk
        my_file = open(file_name,'r')
        return_data  = my_file.readlines()
        my_file.close()
    else:
        return_flag = False
                            
    return (return_flag, return_data)

def get_last(local_file_name, bar):
    # File이 없는경우 대비
    #datetime_old_time   = datetime.datetime(2014, 8, 3, 0,0,0)
    datetime_old_time   = datetime.datetime.now()
    int_old_sec_gap     = 0

    (return_flag, data_list)   = read_file(local_file_name)

    count = 0
    if return_flag == True:
        for line in data_list:
            #item_list값을 밖에서 쓸수없음 bc)file이 없을경우가 잇으므로
            count = int(line)
    else:
        print ('No file and clean Mode')
    #file reading end
    return count

def check_update():
    import datetime
    
    # for both read and save
    local_file_name = 'eukM2.log'
    bar             = '~'
    #~
    
    count =  get_last(local_file_name, bar)
    
    print ("previous value: " + str(count))
    count += 1

        
    my_file = open(local_file_name,"w")    
    my_file.write(str(count))
    my_file.close()    
    
    return int_new_gap


def create_db():
    #여기 두는것은 전체 플로그램을 그냥 시간체크용으로만 쓰는경우도 있어서
    import openpyxl
    from openpyxl import Workbook
        
    old_dict = {}
    if (os.path.exists(db_xls)):
        wb = openpyxl.load_workbook(db_xls)
        sheet_list = wb.get_sheet_names()
        #print (sheet_list)
        sheet = wb.get_sheet_by_name(sheet_list[0])        
        for i in range(1,sheet.max_row + 1):
            key = sheet.cell(row = i, column = 1).value     
            b_debug = True
            b_debug = False        
            if (b_debug == True):
                print (key + ':' + str(sheet.cell(row = i, column = 2).value))
            old_dict[key] = sheet.cell(row = i, column = 2).value
    #endif    

    print ('[creating DB]')
    #
    if os.path.exists(db_xls):
        if os.path.exists('backup_db.xlsx'):
            os.remove('backup_db.xlsx')
        try:    
            os.rename(db_xls,'backup_db.xlsx')
        except :
            print ('e')
            return False
                
    
    file_path_name_list = []
    
    for root, dirs, files in os.walk(mp3_folder_no_endslash):
        for file in files:
            text = os.path.join(root, file)    
            def print_and_append(text):
                try:                
                    print(text)
                except:
                    pass                
                file_path_name_list.append(text)
                      
            if text.find('mp4') != -1 or text.find('mp3') != -1 or text.find('avi') != -1 or text.find('webm') != -1  or text.find('mkv') != -1 or text.find('wav') != -1:
                print_and_append(text)
                
    i_len = len(file_path_name_list)
    print  ('file_path_name_list len:' + str(i_len))
    #for linux
    file_path_name_list.sort();     

    wb = Workbook()
    sheet = wb.active

    row_num = 0;
    for file_path_name in file_path_name_list:
        key = file_path_name	
        row_num += 1
        sheet.cell(row=row_num,column=1).value=key
        
	#for k in old_dict.keys():
	#print 'keys~'
	#print old_dict.keys()
        if key in old_dict.keys():
            sheet.cell(row = row_num,column = 2).value = old_dict[file_path_name]
        else:
            sheet.cell(row = row_num,column = 2).value = '1'
        #print file_path_name+ ':' + str(sheet.cell(row = row_num,column = 2).value)
                
    print ('exist:row_num:' + str(row_num))                
    wb.save(filename = db_xls)   
    return True
    
def create_db_N_create_selected_playlist():
    # create_db 자체를 실패하면 작업취소
    if create_db() == False:        
        return
            
    create_selected_playlist()
    return 

def create_selected_playlist():
    import openpyxl
    from openpyxl import Workbook
    #쓸때 필요
    #pre location for error
    import random
    import shutil
    
    if os.path.isfile(db_xls) == True:                
        wb = openpyxl.load_workbook(db_xls)
        #txt_file = open("eukM2.m3u","w")
        txt_file = open("eukM2.m3u","w",encoding="utf-8")
        
        count = 0
        
        sheet_list = wb.get_sheet_names()
        #print sheet_list
        sheet = wb.get_sheet_by_name(sheet_list[0])
            
        for i in range(1,sheet.max_row + 1):
            i_value = int(sheet.cell(row = i, column = 2).value)                

            print_text = sheet.cell(row = i, column = 1).value + ':' + str(sheet.cell(row = i, column = 2).value)
            random_num = random.randrange(1,i_value + 1)
            if (random_num == 1):                         
                try:
                    print (str(i_value) + '-' + str(random_num) + ':' + print_text + ' Ok')                 
                except:
                    pass
                
                out_txt = sheet.cell(row = i, column = 1).value + '\n'
                if b_windows_or_linux == False:                
                    out_txt = out_txt.replace(mp3_folder_no_endslash + '\\','/my/ongaku/')
                    out_txt = out_txt.replace('\\','/')
                txt_file.write(out_txt)
                count = count + 1
            else:
                try :
                    print (print_text + ' X')                
                except:
                    pass
            
        print (str(count) + '/' + str(sheet.max_row))  
            
        #복사하기전에 close
        txt_file.close()    
        if b_windows_or_linux == False:
            shutil.copy('eukM2.m3u', '\\\\VBOXSVR\\my\\')
            print('복사')        
    else:
        print(db_xls + " 없음")


##main    
'''
main
'''
if __name__ == '__main__':
    work = 10
    work = round(work/60 * 1.1,1)
    #source file 이름찾기?
    import codecs
    source_file = codecs.open('g.py', 'r','utf-8')
    line_list   = source_file.readlines()    
    source_file.close()
    
    count = 0
    for _ in line_list:
        count += 1
        
    # Python Version
    def find_num(text,finding_text):
        index       = text.find(finding_text)
        size        = len(finding_text)
        location    = index + size
        return text[location:location + 1]    

    major = find_num(str(sys.version_info),'major=')
    minor = find_num(str(sys.version_info),'minor=')
    micro = find_num(str(sys.version_info),'micro=')
    print('Git Automation Version: ' + str(work) + 'H ' + str(count) + 'My Python Version: ' +  major + '.' + minor  + '.' + micro)
    #~

    ## global
    b_windows_or_linux = True
    db_xls = 'db.xlsx'    

    # 무슨 모드인지 따라 folder 정함
    if os.path.exists('euk_music_2p3.linux'):
        b_windows_or_linux = False
        print ('Office Mode')
    else:
        b_windows_or_linux = True
        #print ('Home windows Mode')
    ##~ end of global

    #init
    int_new_gap = 0
    check_update()